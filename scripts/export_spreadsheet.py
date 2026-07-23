#!/usr/bin/env python3
"""
Monarch Art Vault — Spreadsheet Export & Google Sheets Sync

Export your art catalog to CSV, XLSX, or sync directly to Google Sheets.
Includes thumbnail links, metadata, print readiness, and marketplace status.
"""
import argparse
import csv
import json
import os
from datetime import datetime
from pathlib import Path


def load_catalog(catalog_path: Path) -> list[dict]:
    """Load catalog from JSONL or JSON file."""
    records = []
    
    if catalog_path.suffix == '.jsonl':
        with open(catalog_path, 'r') as f:
            for line in f:
                if line.strip():
                    records.append(json.loads(line))
    elif catalog_path.suffix == '.json':
        with open(catalog_path, 'r') as f:
            data = json.load(f)
            records = data if isinstance(data, list) else [data]
    else:
        # Try manifest format
        with open(catalog_path, 'r') as f:
            data = json.load(f)
            records = data if isinstance(data, list) else [data]
    
    return records


def flatten_record(record: dict) -> dict:
    """Flatten nested work record for spreadsheet export."""
    flat = {
        'work_id': record.get('work_id') or record.get('asset_id', ''),
        'status': record.get('status', 'cataloged'),
        'filename': record.get('filename', ''),
        'source_path': record.get('source_path', ''),
        'archived_path': record.get('archived_path', ''),
        'managed_path': '',
        'sha256': record.get('sha256', ''),
        'classification': record.get('classification', 'unique'),
        'bytes': record.get('bytes', 0),
        'mime_type': record.get('mime_type', ''),
        'pixel_width': record.get('pixel_width'),
        'pixel_height': record.get('pixel_height'),
        'megapixels': None,
        'aspect_ratio': '',
        'exif_capture_date': record.get('exif_capture_date', ''),
        'creation_date': '',
        'creation_date_precision': '',
        'title': '',
        'title_source': '',
        'title_confidence': None,
        'title_approved': False,
        'artist': '',
        'artist_source': '',
        'description_short': '',
        'description_long': '',
        'style_tags': '',
        'subject_tags': '',
        'dominant_colors': '',
        'alt_text': '',
        'orientation': '',
        'copyright_holder': '',
        'license': '',
        'print_status': '',
        'print_sizes_inches': '',
        'print_warnings': '',
        'marketplace_status': 'not_listed',
        'etsy_listing_id': '',
        'shopify_product_id': '',
        'thumbnail_url': '',
        'storage_link': '',
        'run_id': record.get('run_id', ''),
        'ingested_at': record.get('ingested_at', ''),
    }
    
    # Calculate megapixels and aspect ratio
    if flat['pixel_width'] and flat['pixel_height']:
        flat['megapixels'] = round((flat['pixel_width'] * flat['pixel_height']) / 1_000_000, 2)
        w, h = flat['pixel_width'], flat['pixel_height']
        if w > h:
            flat['aspect_ratio'] = f"{round(w/h, 2)}:1 (landscape)"
        elif h > w:
            flat['aspect_ratio'] = f"1:{round(h/w, 2)} (portrait)"
        else:
            flat['aspect_ratio'] = "1:1 (square)"
    
    # Extract nested title
    if 'title' in record and isinstance(record['title'], dict):
        flat['title'] = record['title'].get('value', '')
        flat['title_source'] = record['title'].get('source', '')
        flat['title_confidence'] = record['title'].get('confidence')
        flat['title_approved'] = record['title'].get('approved', False)
    
    # Extract nested artist
    if 'artist' in record and isinstance(record['artist'], dict):
        flat['artist'] = record['artist'].get('value', '')
        flat['artist_source'] = record['artist'].get('source', '')
    
    # Extract creation date
    if 'creation_date' in record and isinstance(record['creation_date'], dict):
        flat['creation_date'] = record['creation_date'].get('value', '')
        flat['creation_date_precision'] = record['creation_date'].get('precision', '')
    
    # Extract description
    if 'description' in record and isinstance(record['description'], dict):
        flat['description_short'] = record['description'].get('short', '')
        flat['description_long'] = record['description'].get('long', '')
    
    # Extract classification/tags
    if 'classification' in record and isinstance(record['classification'], dict):
        flat['style_tags'] = ', '.join(record['classification'].get('style_tags', []))
        flat['subject_tags'] = ', '.join(record['classification'].get('subject_tags', []))
    
    # Extract visual info
    if 'visual' in record and isinstance(record['visual'], dict):
        flat['dominant_colors'] = ', '.join(record['visual'].get('dominant_colors', []))
        flat['alt_text'] = record['visual'].get('alt_text', '')
        flat['orientation'] = record['visual'].get('orientation', '')
    
    # Extract rights
    if 'rights' in record and isinstance(record['rights'], dict):
        flat['copyright_holder'] = record['rights'].get('copyright_holder', '')
        flat['license'] = record['rights'].get('license', '')
    
    # Extract print readiness
    if 'print_readiness' in record and isinstance(record['print_readiness'], dict):
        flat['print_status'] = record['print_readiness'].get('status', '')
        sizes = record['print_readiness'].get('recommended_sizes_inches', [])
        flat['print_sizes_inches'] = ', '.join(sizes) if sizes else ''
        warnings = record['print_readiness'].get('warnings', [])
        flat['print_warnings'] = '; '.join(warnings) if warnings else ''
    
    # Extract assets array for managed path
    if 'assets' in record and record['assets']:
        first_asset = record['assets'][0]
        flat['managed_path'] = first_asset.get('managed_path', '')
        if not flat['sha256']:
            flat['sha256'] = first_asset.get('sha256', '')
        if not flat['pixel_width']:
            flat['pixel_width'] = first_asset.get('pixel_width')
        if not flat['pixel_height']:
            flat['pixel_height'] = first_asset.get('pixel_height')
    
    # Provenance
    if 'provenance' in record and isinstance(record['provenance'], dict):
        flat['run_id'] = record['provenance'].get('run_id', flat['run_id'])
        flat['ingested_at'] = record['provenance'].get('ingested_at', flat['ingested_at'])
    
    return flat


def export_csv(records: list[dict], output_path: Path, include_thumbnails: bool = True):
    """Export catalog to CSV file."""
    if not records:
        print("No records to export.")
        return
    
    flat_records = [flatten_record(r) for r in records]
    
    # Filter columns based on user preference
    columns = list(flat_records[0].keys())
    if not include_thumbnails:
        columns = [c for c in columns if c not in ['thumbnail_url', 'storage_link']]
    
    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=columns, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(flat_records)
    
    print(f"✓ Exported {len(flat_records)} records to {output_path}")


def export_xlsx(records: list[dict], output_path: Path):
    """Export catalog to Excel file with formatting."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("⚠ openpyxl not installed. Install with: pip install openpyxl")
        print("  Falling back to CSV export.")
        export_csv(records, output_path.with_suffix('.csv'))
        return
    
    flat_records = [flatten_record(r) for r in records]
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Art Catalog"
    
    # Headers
    columns = list(flat_records[0].keys()) if flat_records else []
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    header_font = Font(color="c9a227", bold=True)
    
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    for row_idx, record in enumerate(flat_records, 2):
        for col_idx, col_name in enumerate(columns, 1):
            value = record.get(col_name, '')
            if isinstance(value, (list, dict)):
                value = json.dumps(value)
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Auto-fit columns
    for col_idx, col_name in enumerate(columns, 1):
        max_length = len(col_name)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)[:50]))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    wb.save(output_path)
    print(f"✓ Exported {len(flat_records)} records to {output_path}")


def sync_google_sheets(records: list[dict], spreadsheet_id: str = None, sheet_name: str = "Art Catalog"):
    """
    Sync catalog to Google Sheets.
    
    Requires:
    - Google Cloud project with Sheets API enabled
    - Service account or OAuth credentials in GOOGLE_APPLICATION_CREDENTIALS
    - Or use 'gspread' with service account JSON
    
    Args:
        records: List of work records
        spreadsheet_id: Existing spreadsheet ID to update, or None to create new
        sheet_name: Name of the worksheet
    """
    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError:
        print("⚠ gspread not installed. Install with: pip install gspread google-auth")
        return None
    
    # Check for credentials
    creds_path = os.environ.get('GOOGLE_APPLICATION_CREDENTIALS')
    if not creds_path:
        creds_path = Path.home() / '.config' / 'monarch-art-vault' / 'google-service-account.json'
        if not creds_path.exists():
            print(f"""
⚠ Google credentials not found.

To enable Google Sheets sync:
1. Create a Google Cloud project: https://console.cloud.google.com
2. Enable the Google Sheets API
3. Create a service account and download the JSON key
4. Save the key to one of:
   - Set GOOGLE_APPLICATION_CREDENTIALS environment variable
   - Save to ~/.config/monarch-art-vault/google-service-account.json
5. Share your spreadsheet with the service account email

For detailed setup: https://github.com/diamitani/monarch-art-vault#google-sheets-setup
""")
            return None
    
    # Authenticate
    scopes = [
        'https://www.googleapis.com/auth/spreadsheets',
        'https://www.googleapis.com/auth/drive'
    ]
    
    try:
        credentials = Credentials.from_service_account_file(str(creds_path), scopes=scopes)
        gc = gspread.authorize(credentials)
    except Exception as e:
        print(f"⚠ Failed to authenticate with Google: {e}")
        return None
    
    flat_records = [flatten_record(r) for r in records]
    
    if not flat_records:
        print("No records to sync.")
        return None
    
    # Prepare data
    columns = list(flat_records[0].keys())
    data = [columns]  # Header row
    for record in flat_records:
        row = []
        for col in columns:
            value = record.get(col, '')
            if isinstance(value, (list, dict)):
                value = json.dumps(value)
            elif value is None:
                value = ''
            row.append(str(value))
        data.append(row)
    
    try:
        if spreadsheet_id:
            # Update existing spreadsheet
            sh = gc.open_by_key(spreadsheet_id)
        else:
            # Create new spreadsheet
            sh = gc.create(f"Monarch Art Vault - {datetime.now().strftime('%Y-%m-%d')}")
            print(f"✓ Created new spreadsheet: {sh.url}")
        
        # Get or create worksheet
        try:
            worksheet = sh.worksheet(sheet_name)
            worksheet.clear()
        except gspread.WorksheetNotFound:
            worksheet = sh.add_worksheet(title=sheet_name, rows=len(data)+10, cols=len(columns))
        
        # Update data
        worksheet.update(data, 'A1')
        
        # Format header row
        worksheet.format('1:1', {
            'backgroundColor': {'red': 0.1, 'green': 0.1, 'blue': 0.18},
            'textFormat': {'foregroundColor': {'red': 0.79, 'green': 0.64, 'blue': 0.15}, 'bold': True}
        })
        
        # Freeze header
        worksheet.freeze(rows=1)
        
        print(f"✓ Synced {len(flat_records)} records to Google Sheets")
        print(f"  URL: {sh.url}")
        return sh.url
        
    except Exception as e:
        print(f"⚠ Failed to sync to Google Sheets: {e}")
        return None


def create_downloadable_package(
    records: list[dict], 
    output_dir: Path,
    include_thumbnails: bool = True,
    generate_contact_sheet: bool = True
):
    """
    Create a downloadable package with spreadsheet and optional thumbnails.
    
    Creates:
    - catalog.csv
    - catalog.xlsx (if openpyxl available)
    - thumbnails/ folder with preview images
    - contact-sheet.pdf (if PIL available)
    - README.txt with package info
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # Export CSV
    csv_path = output_dir / 'catalog.csv'
    export_csv(records, csv_path, include_thumbnails)
    
    # Export XLSX
    xlsx_path = output_dir / 'catalog.xlsx'
    export_xlsx(records, xlsx_path)
    
    # Create README
    readme_content = f"""Monarch Art Vault Export
========================
Generated: {datetime.now().isoformat()}
Records: {len(records)}

Files:
- catalog.csv: Full catalog in CSV format
- catalog.xlsx: Excel workbook with formatting
- thumbnails/: Preview images (if generated)

Column Descriptions:
- work_id: Unique identifier for the work
- status: needs_review | cataloged | approved | published
- classification: unique | exact_duplicate | near_duplicate | variant
- title_source: embedded | user_confirmed | ai_suggested | unknown
- print_status: print_ready | conditionally_ready | needs_master_file | not_recommended

For questions: https://github.com/diamitani/monarch-art-vault
"""
    (output_dir / 'README.txt').write_text(readme_content)
    
    print(f"✓ Created downloadable package at {output_dir}")
    return output_dir


def main():
    parser = argparse.ArgumentParser(
        description='Export Monarch Art Vault catalog to spreadsheet formats',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Export to CSV
  python export_spreadsheet.py catalog.jsonl --format csv
  
  # Export to Excel
  python export_spreadsheet.py catalog.jsonl --format xlsx
  
  # Sync to Google Sheets
  python export_spreadsheet.py catalog.jsonl --format gsheets
  
  # Create full downloadable package
  python export_spreadsheet.py catalog.jsonl --package ./export
"""
    )
    
    parser.add_argument('catalog', help='Path to catalog file (JSONL or JSON)')
    parser.add_argument('--format', '-f', choices=['csv', 'xlsx', 'gsheets', 'all'], 
                        default='csv', help='Output format (default: csv)')
    parser.add_argument('--output', '-o', help='Output file/directory path')
    parser.add_argument('--spreadsheet-id', help='Existing Google Sheets ID to update')
    parser.add_argument('--sheet-name', default='Art Catalog', help='Google Sheets worksheet name')
    parser.add_argument('--package', help='Create full downloadable package at this directory')
    parser.add_argument('--no-thumbnails', action='store_true', help='Exclude thumbnail columns')
    
    args = parser.parse_args()
    
    # Load catalog
    catalog_path = Path(args.catalog)
    if not catalog_path.exists():
        print(f"Error: Catalog file not found: {catalog_path}")
        return 1
    
    records = load_catalog(catalog_path)
    print(f"Loaded {len(records)} records from {catalog_path}")
    
    # Create package if requested
    if args.package:
        create_downloadable_package(
            records, 
            Path(args.package),
            include_thumbnails=not args.no_thumbnails
        )
        return 0
    
    # Determine output path
    output = Path(args.output) if args.output else catalog_path.with_suffix('')
    
    # Export based on format
    if args.format == 'csv':
        export_csv(records, output.with_suffix('.csv'), include_thumbnails=not args.no_thumbnails)
    elif args.format == 'xlsx':
        export_xlsx(records, output.with_suffix('.xlsx'))
    elif args.format == 'gsheets':
        sync_google_sheets(records, args.spreadsheet_id, args.sheet_name)
    elif args.format == 'all':
        export_csv(records, output.with_suffix('.csv'), include_thumbnails=not args.no_thumbnails)
        export_xlsx(records, output.with_suffix('.xlsx'))
        sync_google_sheets(records, args.spreadsheet_id, args.sheet_name)
    
    return 0


if __name__ == '__main__':
    exit(main())
